#COMP231-002 Team4
#301296260 Chung Ching Lai
#301279252 Man Fai Kwan
#301282811 Man Chun Lam
#301272631 Fatima Khalid
#3xxxxxxxx yyyyyyyyyyyyyyy
#3xxxxxxxx yyyyyyyyyyyyyyy
#Date Last Modified: <2024-04-12>
#Program Title: Automatic Data Logger
#Version : 00
#Revision History:
# V00 - 2024-04-12 First release
import cv2
from tkinter import *
import tkinter.ttk as ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import numpy as np
import openpyxl
from sklearn.neighbors import KNeighborsClassifier
import os
from datetime import datetime ,timedelta

version="V00"
demoMode=True
demoVideo='shorttest.mp4'
###### AI parameters ###########################
nSize   = 50     #pixel of training images
displayTime = 300      #eg 500 = 0.5 sec , 0=inf
################################################


class AppLogin:
    # Login Page
    def __init__(self, master):
        self.master = master
        self.master.title('Login')
        self.master.geometry("200x120+500+300")

        self.username = None
        self.password = None

        # Username
        self.label_username = Label(master, text="Username")
        self.label_username.grid(row=0, column=0, sticky='e')
        self.entry_username = Entry(master)
        self.entry_username.grid(row=0, column=1)

        # Password
        self.label_password = Label(master, text="Password")
        self.label_password.grid(row=1, column=0, sticky='e')
        self.entry_password = Entry(master, show="*")
        self.entry_password.grid(row=1, column=1)

        # Login button
        self.login_button = Button(master, text="Login", command=self.check_login)
        self.login_button.grid(row=2, column=0, columnspan=2)

        # Message label
        self.message_label = Label(master, text="Username: admin", fg="blue")
        self.message_label.grid(row=3, column=0, columnspan=2, sticky='w')
        self.message_label = Label(master, text="Password: password", fg="blue")
        self.message_label.grid(row=4, column=0, columnspan=2, sticky='w')

        master.protocol("WM_DELETE_WINDOW", self.on_close)
        # This makes the window wait until the login is successful or the window is closed
        self.master.mainloop()

    def check_login(self):
        # Hard-coded credentials for demonstration; replace with your authentication method
        username = self.entry_username.get()
        password = self.entry_password.get()

        if username == "admin" and password == "password":
            messagebox.showinfo("Login Info", "Welcome!")
            self.master.destroy()  # Close the login window
        else:
            messagebox.showerror("Login Info", "Incorrect username or password")
    
    def on_close(self):
        # This method is called when the window is closed
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.master.destroy()
            os._exit(0)
class MyVideoCapture:
    # This class is used to access Camera
    #
    # Methods:
    # - get_frame(): get frame data from Camera
    def __init__(self, video_source=0):
        # Open the video source
        if demoMode:
            self.vid=cv2.VideoCapture(demoVideo)
        else:
            self.vid = cv2.VideoCapture(video_source)
        
        if not self.vid.isOpened():
            raise ValueError("Unable to open video source", video_source)
        
        self.vid.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        self.vid.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        # Get video source width and height
        self.width = self.vid.get(cv2.CAP_PROP_FRAME_WIDTH)
        self.height = self.vid.get(cv2.CAP_PROP_FRAME_HEIGHT)
 
    # Release the video source when the object is destroyed
    def __del__(self):
        # release Camera source
        # Return: None
        if self.vid.isOpened():
            self.vid.release()
    
    def get_frame(self):
        # get frame data
        # return: frame data
        if self.vid.isOpened():
            ret, frame = self.vid.read()

            if not ret and demoMode:
                self.vid = cv2.VideoCapture(demoVideo)
                ret, frame = self.vid.read()
            if ret:
                return (ret, frame)
            else:
                return (ret, None)
        else:
            return (ret, None)


class App:
    # This class is used to display UI for monitoring and operation
    def __init__(self, window, window_title, video_source=0):
        self.window = window
        self.window.title(window_title)
        self.window.geometry("+40+40")
        self.video_source = video_source
        self.xLimit=639
        self.yLimit=479
        self.currentX=0
        self.currentY=0
        self.line1=None
        self.line2=None
        self.overLine1=None
        self.overLine2=None
        self.overGrid=[]
        self.startX=0
        self.startY=0
        self.setArea=False
        self.areaObject=[]
        self.nameObject=[]
        self.areaName=[]
        self.areaNumber=0
        self.areaCoord=[]
        self.temporaryArea=None
        self.aiModel=None
        self.aiResult=[]
        self.aiResultDict={}
        self.lastLogTime=datetime(2000, 1, 1, 00, 00, 00)
        self.logInterval=0
        self.rawImgName = []
        self.rawLabel = []
        # open video source
        self.vid = MyVideoCapture(video_source)
 
        ######## UI Elements #############
        # Create a canvas that can fit the above video source size
        self.canvas = Canvas(window, width = self.vid.width, height = self.vid.height,relief='ridge',bd=1, highlightthickness=1,)
        self.btnCapture = Button(window,text='CAPTURE', command=self.save)
        self.vX = StringVar()
        self.vX.set("X=")
        self.vY = StringVar()
        self.vY.set("Y=")
        self.xLabel=Label(self.window,textvariable=self.vX)
        self.yLabel=Label(self.window,textvariable=self.vY)
        self.vOverLine=StringVar()
        self.vOverLine.set('NONE')
        self.rd1=Radiobutton(self.window,text='Center Line',variable=self.vOverLine,value='CENT',command=self.changeOverlayLine)
        self.rd2=Radiobutton(self.window,text='Grid Line',variable=self.vOverLine,value='GRID',command=self.changeOverlayLine)
        self.rd3=Radiobutton(self.window,text='None',variable=self.vOverLine,value='NONE',command=self.changeOverlayLine)
        self.areaNameLabel=Label(self.window,text="Area Name:")
        self.areaNameEntry=Entry(self.window,width=10)
        self.areaNameEntry.insert(END,'')
        self.btnDeleteArea = Button(window,text='Delete Area by name', command=self.deleteArea)

        self.selectArea_cb=ttk.Combobox(self.window, width=5)
        self.selectArea_cb['values']=self.areaName
        self.selectArea_cb['state']='readonly'
        self.captureFileEntry=Entry(self.window,width=15)
        self.captureFileEntry.insert(END,'')

        self.btnOpenSetting = Button(self.window,text='Open Setting', command=self.openSetting)
        self.btnSaveSetting = Button(self.window,text='Save Setting', command=self.saveSetting)
        self.btnLoadSetting = Button(self.window,text='Load Setting', command=self.loadSetting)
        self.btnResetSetting = Button(self.window,text='Reset Setting', command=self.resetSetting)
        self.settingFileEntry=Entry(self.window,width=30)
        self.settingFileEntry.insert(END,'setting')

        self.var_checkB1=BooleanVar()
        self.B1Check=Checkbutton(window, text='AI',variable=self.var_checkB1,command=self.trainAI)
        self.var_showTrainImg=BooleanVar()
        self.showTrainImg=Checkbutton(window, text='Train Img.', variable=self.var_showTrainImg)
        self.var_showTestImg=BooleanVar()
        self.showTestImg=Checkbutton(window, text='Test Img.', variable=self.var_showTestImg)

        self.aiColor=StringVar()
        self.aiColor.set('COLOR')
        self.rd4=Radiobutton(self.window,text='COLOR',variable=self.aiColor,value='COLOR',command=self.changeAIColor)
        self.rd5=Radiobutton(self.window,text='GRAY',variable=self.aiColor,value='GRAY',command=self.changeAIColor)
        self.rd6=Radiobutton(self.window,text='BW',variable=self.aiColor,value='BW',command=self.changeAIColor)

        self.hourLabel=Label(self.window,text="Hours")
        self.minuteLabel=Label(self.window,text="Minutes")
        self.secondLabel=Label(self.window,text="Seconds")
        self.hour_cb=ttk.Combobox(self.window, width=5)
        self.hour_cb['values']=list(range(1000))
        self.hour_cb['state']='readonly'
        self.hour_cb.current(0)
        self.minute_cb=ttk.Combobox(self.window, width=5)
        self.minute_cb['values']=list(range(60))
        self.minute_cb['state']='readonly'
        self.minute_cb.current(0)
        self.second_cb=ttk.Combobox(self.window, width=5)
        self.second_cb['values']=list(range(60))
        self.second_cb['state']='readonly'
        self.second_cb.current(0)
        self.logFormatLabel=Label(self.window,text="Log Format:")
        self.btnLogFile = Button(self.window,text='Open logfile:', command=self.openLogFile)
        self.logFileEntry=Entry(self.window,width=58)
        self.logFileEntry.insert(END,'log.csv')
        self.logFormatEntry=Entry(self.window,width=58)
        self.logFormatEntry.insert(END,'[$DATE],[$TIME],[1][2]:[3][4]')

        ######### GRID SETTING ###########
        self.canvas.grid(row=0,column=0,rowspan=1,columnspan=20,sticky=NW)
        self.xLabel.grid(row=1,column=0,sticky=W)
        self.yLabel.grid(row=1,column=1,sticky=W)
        self.rd1.grid(row=1,column=2,sticky=W)
        self.rd2.grid(row=1,column=3,sticky=W)
        self.rd3.grid(row=1,column=4,sticky=W)
        self.areaNameLabel.grid(row=1,column=5,sticky=W)
        self.areaNameEntry.grid(row=1,column=6,sticky=W)
        self.btnDeleteArea.grid(row=1,column=7,sticky=W)
        self.btnCapture.grid(row=2, column=0, sticky=W)
        self.selectArea_cb.grid(row=2,column=1,sticky=W)
        self.captureFileEntry.grid(row=2,column=2, columnspan=4, sticky=W)

        self.btnOpenSetting.grid(row=2, column=3, sticky=W)
        self.btnSaveSetting.grid(row=2, column=4,sticky=W)
        self.btnLoadSetting.grid(row=2, column=5,sticky=W)
        self.settingFileEntry.grid(row=2, column=6, columnspan=2, sticky=W)
        
        
        self.btnResetSetting.grid(row=3, column=0,columnspan=2, sticky=W)
        self.rd4.grid(row=3,column=2,sticky=W)
        self.rd5.grid(row=3,column=3,sticky=W)
        self.rd6.grid(row=3,column=4,sticky=W)
        self.showTrainImg.grid(row=3,column=5,sticky=W)
        self.showTestImg.grid(row=3,column=6,sticky=W)
        self.B1Check.grid(row=3,column=7,sticky=W)

        self.hourLabel.grid(row=4,column=0,sticky=W)
        self.minuteLabel.grid(row=4,column=1,sticky=W)
        self.secondLabel.grid(row=4,column=2,sticky=W)
        self.hour_cb.grid(row=5,column=0,sticky=W)
        self.minute_cb.grid(row=5,column=1,sticky=W)
        self.second_cb.grid(row=5,column=2,sticky=W)
        self.logFormatLabel.grid(row=4,column=3,sticky=W)
        self.btnLogFile.grid(row=5, column=3, sticky=W)
        self.logFileEntry.grid(row=5,column=4,columnspan=10,sticky=W)
        self.logFormatEntry.grid(row=4,column=4,columnspan=10,sticky=W)

        ######### BIND #########
        self.canvas.bind("<ButtonPress-1>", self.on_button_press)
        self.canvas.bind("<Motion>", self.on_move_press)
        self.canvas.bind('<ButtonPress-3>', self.b3button)
        self.hour_cb.bind('<<ComboboxSelected>>', self.setInterval)
        self.minute_cb.bind('<<ComboboxSelected>>', self.setInterval)
        self.second_cb.bind('<<ComboboxSelected>>', self.setInterval)

        self.delay = 15
        self.update()
        self.window.mainloop()

    def update(self):
        # redraw a frame from the video source
        # redraw screen objects
        # return: none
        ret, frame = self.vid.get_frame()
        
        if ret:
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            self.photo = ImageTk.PhotoImage(image = Image.fromarray(frame))
            self.canvas.create_image(0, 0, image = self.photo, anchor = 'nw')
            # redraw cross line objects
            if self.line1:
                self.canvas.itemconfig(self.line1, state='normal')
                self.canvas.itemconfig(self.line2, state='normal')
                self.canvas.tag_raise(self.line1)
                self.canvas.tag_raise(self.line2)
            # redraw overlay lines objects
            if self.vOverLine.get()!='NONE':
                if self.vOverLine.get()=='CENT':
                    self.canvas.itemconfig(self.overLine1, state='normal')
                    self.canvas.itemconfig(self.overLine2, state='normal')
                    self.canvas.tag_raise(self.overLine1)
                    self.canvas.tag_raise(self.overLine2)
                elif self.vOverLine.get()=='GRID':
                    for gridLine in self.overGrid:
                        self.canvas.itemconfig(gridLine, state='normal')
                        self.canvas.tag_raise(gridLine)
            # redraw rectangle on the fly
            if self.setArea:
                self.temporaryArea=self.canvas.create_rectangle(self.startX, self.startY, self.currentX, self.currentY,fill="",outline= "GREEN")
            # redraw area rectangles
            if self.areaObject:
                for area in self.areaObject:
                    self.canvas.tag_raise(area)
                for name in self.nameObject:
                    self.canvas.tag_raise(name)
            # redraw AI objects
            if self.var_checkB1.get():
                for result in self.aiResult:
                    if result:
                        self.canvas.tag_raise(result)
            
        self.window.after(self.delay, self.update)

    # cross lines to indicate the location of mouse pointer - David
    def on_move_press(self, event):
        self.currentX, self.currentY = (event.x, event.y)
        if self.currentX>self.xLimit:
            self.currentX=self.xLimit
        if self.currentY>self.yLimit:
            self.currentY=self.yLimit
        self.vX.set("X="+str(self.currentX))
        self.vY.set("Y="+str(self.currentY))
        if self.line1:
            self.canvas.delete(self.line1)
        if self.line2:
            self.canvas.delete(self.line2)
        self.line1=self.canvas.create_line(0, self.currentY, self.xLimit , self.currentY,fill='lime green',state='hidden')
        self.line2=self.canvas.create_line(self.currentX, 0, self.currentX, self.yLimit,fill='lime green',state='hidden')

# Code to draw a rectangle when left button of the mouse is clicked - Fatima 
    def on_button_press(self, event):
        self.setArea=not self.setArea
        if self.setArea:
            self.startX,self.startY=(event.x,event.y)
        else:
            self.areaObject.append(self.canvas.create_rectangle(self.startX, self.startY, self.currentX, self.currentY,fill="",outline= "LIME"))
            if self.areaNameEntry.get() and not self.areaNameEntry.get().isspace():
                self.nameObject.append(self.canvas.create_text(self.startX+3, self.startY-10, text=self.areaNameEntry.get(), fill="LIME", font=('Helvetica 15'), anchor='w'))
                self.areaName.append(self.areaNameEntry.get())
            else:
                self.areaNumber=self.areaNumber+1
                self.nameObject.append(self.canvas.create_text(self.startX+3, self.startY-10, text=str(self.areaNumber), fill="LIME", font=('Helvetica 15'), anchor='w'))
                self.areaName.append(str(self.areaNumber))
            self.areaCoord.append([min(self.startX,self.currentX), min(self.startY,self.currentY), max(self.currentX,self.startX), max(self.currentY,self.startY)])
            self.selectArea_cb['values']=self.areaName
            self.selectArea_cb['state']='readonly'
            self.aiResult.append(None)
            print(self.areaName[-1]+'@',self.canvas.coords(self.areaObject[-1]))
            print(self.areaName[-1]+'#',self.areaCoord[-1])

        
    def b3button(self,event):
        # Cancel drawing rectangle
        # return: none
        self.setArea=False
        self.temporaryArea=None

    def save(self):
        # Capture area screenshots and save into file
        # return: none
        if self.selectArea_cb.get():
            response = False
            if os.path.exists('./reference/'+self.captureFileEntry.get()+'.bmp'):
                    response = messagebox.askyesno("Question","Overwrite the existing file?")
            else:
                    response=True

            if response:
                for index, name in enumerate(self.areaName):
                    if name==self.selectArea_cb.get():
                        startCol,startRow,endCol,endRow=self.areaCoord[index]
                        ret, frame = self.vid.get_frame()
                        crop = frame[int(startRow):int(endRow),int(startCol):int(endCol)]
                        if self.captureFileEntry.get() and not self.captureFileEntry.get().isspace():
                            cv2.imwrite('./reference/'+self.captureFileEntry.get()+".bmp",crop)
                        break

    # save setting file function - David
    def saveSetting(self):
        # save setting file
        # return: none
        saveFile=False
        if self.settingFileEntry.get() and (not self.settingFileEntry.get().isspace()):
            if os.path.exists(self.settingFileEntry.get()+'.xlsx'):
                response = messagebox.askyesno("Question","Overwrite the existing file?")
                if response:
                    saveFile=True
            else:
                saveFile=True

            if saveFile:
                wb=openpyxl.Workbook()
                worksheet=wb.active
                worksheet.title='Sheet1'
                row = 1
                for index, name in enumerate(self.areaName):
                    x1,y1,x2,y2 = self.areaCoord[index]
                    worksheet.cell(row = row, column = 1).value = 'AREA'
                    worksheet.cell(row = row, column = 2).value = name
                    worksheet.cell(row = row, column = 3).value = x1
                    worksheet.cell(row = row, column = 4).value = y1
                    worksheet.cell(row = row, column = 5).value = x2
                    worksheet.cell(row = row, column = 6).value = y2
                    row = row + 1
                for index, imgName in enumerate(self.rawImgName):
                    worksheet.cell(row = row, column = 1).value = 'MAPPING'
                    worksheet.cell(row = row, column = 2).value = imgName
                    worksheet.cell(row = row, column = 3).value = self.rawLabel[index]
                    row = row + 1
                worksheet.cell(row = row, column = 1).value = 'AICOLOR'
                worksheet.cell(row = row, column = 2).value = self.aiColor.get()
                row = row + 1
                worksheet.cell(row = row, column = 1).value = 'INTERVAL'
                worksheet.cell(row = row, column = 2).value = self.hour_cb.get()
                worksheet.cell(row = row, column = 3).value = self.minute_cb.get()
                worksheet.cell(row = row, column = 4).value = self.second_cb.get()
                row = row + 1
                worksheet.cell(row = row, column = 1).value = 'LOGFILE'
                worksheet.cell(row = row, column = 2).value = self.logFileEntry.get()
                row = row + 1
                worksheet.cell(row = row, column = 1).value = 'LOGFORMAT'
                worksheet.cell(row = row, column = 2).value = self.logFormatEntry.get()
                wb.save(filename=self.settingFileEntry.get()+".xlsx")

    def loadSetting(self):
        # Load setting file
        # return: none
        if self.settingFileEntry.get() and (not self.settingFileEntry.get().isspace()):
            if os.path.exists(self.settingFileEntry.get()+'.xlsx'):
                if self.areaObject or self.rawImgName:
                    response = messagebox.askyesno("Question","This action will merge the current setting with setting in file, continues?")
                else:
                    response = True
                
                if response:
                    self.var_checkB1.set(False)
                    wb = openpyxl.load_workbook(filename=self.settingFileEntry.get()+".xlsx",data_only=True)
                    sheet = wb['Sheet1']
                    setting = []
                    for value in sheet.iter_rows(values_only=True):
                        setting.append(value)
                    for data in setting:
                        if data[0]=='AICOLOR':
                            self.aiColor.set(data[1])
                        if data[0]=='AREA':
                            self.areaName.append(data[1])
                            self.nameObject.append(self.canvas.create_text(data[2]+3, data[3]-10, text=str(data[1]), fill="LIME", font=('Helvetica 15'), anchor='w'))
                            self.areaNumber = self.areaNumber + 1
                            self.areaCoord.append([data[2], data[3], data[4], data[5]])
                            self.areaObject.append(self.canvas.create_rectangle(data[2], data[3], data[4], data[5],fill="",outline= "LIME"))
                            self.aiResult.append(None)
                        if data[0]=='MAPPING':
                            self.rawImgName.append(data[1])
                            self.rawLabel.append(str(data[2]))
                        if data[0]=='INTERVAL':
                            self.hour_cb.current(int(data[1]))
                            self.minute_cb.current(int(data[2]))
                            self.second_cb.current(int(data[3]))
                            self.logInterval = int(self.hour_cb.get())*60*60 + int(self.minute_cb.get())*60 + int(self.second_cb.get())
                        if data[0]=='LOGFILE':
                            self.logFileEntry.delete(0,END)
                            self.logFileEntry.insert(END,data[1])
                        if data[0]=='LOGFORMAT':
                            self.logFormatEntry.delete(0,END)
                            self.logFormatEntry.insert(END,data[1])
                    self.selectArea_cb['values']=self.areaName
                    self.selectArea_cb['state']='readonly'
                    
                
        
    def openSetting(self):
        # open setting file
        # retrun: none
        if not os.path.exists(self.settingFileEntry.get()+'.xlsx'):
            response = messagebox.showinfo("Info","File does not exist!")
        else:
            os.system("start EXCEL.EXE "+ self.settingFileEntry.get()+".xlsx")
    
    def resetSetting(self):
        # Clear all setting and reset to default states
        # return: none
        self.var_checkB1.set(False)
        self.areaObject=[]
        self.nameObject=[]
        self.areaName=[]
        self.areaNumber=0
        self.areaCoord=[]
        self.aiResult=[]
        self.rawImgName=[]
        self.rawLabel=[]
        self.selectArea_cb['values']=self.areaName
        self.selectArea_cb['state']='readonly'
        self.hour_cb.current(0)
        self.minute_cb.current(0)
        self.second_cb.current(0)
        self.logFileEntry.delete(0,END)
        self.logFileEntry.insert(END,'log.csv')
        self.logFormatEntry.delete(0,END)
        self.logFormatEntry.insert(END,'[$DATE],[$TIME],[1][2]:[3][4]')

    def setInterval(self,event):
        # set interval time
        # return: none
        self.logInterval = int(self.hour_cb.get())*60*60 + int(self.minute_cb.get())*60 + int(self.second_cb.get())

    def saveLogFile(self):
        # save log file insert value according to format setting
        # return: none
        if self.logFileEntry.get() and not self.logFileEntry.get().isspace():
            template = self.logFormatEntry.get()
                        
            result=template
            for k,v in self.aiResultDict.items():
                result = result.replace('[' + k + ']', str(v))
            with open(self.logFileEntry.get(), 'a', newline='') as file:
                file.write(result + '\n')

    def openLogFile(self):
        # open log file
        # return : none
        if not os.path.exists(self.logFileEntry.get()):
            response = messagebox.showinfo("Info","File does not exist!")
        else:
            os.system("start NOTEPAD.EXE "+ self.logFileEntry.get())


    #delete area - David
    def deleteArea(self):
        # delete an Area
        # return: none
        scanAll=False
        while(not scanAll):
            scanAll=True
            for index, name in enumerate(self.areaName):
                if name==self.areaNameEntry.get():
                    self.nameObject.pop(index)
                    self.areaObject.pop(index)
                    self.areaName.pop(index)
                    self.areaCoord.pop(index)
                    self.aiResult.pop(index)
                    scanAll=False
                    break
        self.selectArea_cb['values']=self.areaName
        self.selectArea_cb['state']='readonly'
  
    def changeOverlayLine(self):
        # Change overlay Lines
        # return: none
        if self.vOverLine.get()=='CENT':
            self.overLine1=self.canvas.create_line(0, 240, self.xLimit , 240,fill='GREEN',state='hidden')
            self.overLine2=self.canvas.create_line(320, 0, 320 , self.yLimit,fill='GREEN',state='hidden')
        elif self.vOverLine.get()=='GRID':
            for i in range(0,self.xLimit, 80):
                self.overGrid.append(self.canvas.create_line(i, 0, i , 479,fill='GREEN',state='hidden'))
            for j in range(0,self.yLimit, 80):
                self.overGrid.append(self.canvas.create_line(0, j, 639 , j,fill='GREEN',state='hidden'))


    def changeAIColor(self):
        # Select color for AI traing and recognization
        # return: none
        self.var_checkB1.set(False)
    
    def trainAI(self):
        # Training AI and create AI model
        # return : none
        if self.var_checkB1.get() and self.rawImgName:
            print('Start AI')
            trainImg=[]
            trainLabel=[]
            for index, name in enumerate(self.rawImgName):
                    trainImg.append(prepareImageF('./reference/'+name, nSize,display=self.var_showTrainImg.get(),color=self.aiColor.get()))
                    trainLabel.append(str(self.rawLabel[index]))
            for j in range(9):
                for index, name in enumerate(self.rawImgName):
                    trainImg.append(prepareImageF('./reference/'+name, nSize,display=False,color=self.aiColor.get()))
                    trainLabel.append(str(self.rawLabel[index]))
            
            # Define the model architecture
            self.aiModel= KNeighborsClassifier()
            
            
            # training
            self.aiModel.fit(trainImg, trainLabel)

            # start prediction
            self.runAI()
        else:
            print('AI stopped')
            if not self.rawImgName:
                response = messagebox.showinfo("Info","No training data!")
            self.var_checkB1.set(False)

    def runAI(self):
        # Consume AI model to recognize digits
        # return: none
        if self.var_checkB1.get():
            if self.areaObject:
                ret,frame=self.vid.get_frame()
                counter = 0
                now = datetime.now()
                self.aiResultDict['$DATE']=now.date()
                self.aiResultDict['$TIME']=now.time()
                for area in self.areaCoord:
                    startCol,startRow,endCol,endRow=area
                    
                    crop = frame[int(startRow):int(endRow),int(startCol):int(endCol)]
                    test_x=[]
                    test_x.append(prepareImageM(crop, nSize,display=self.var_showTestImg.get(),color=self.aiColor.get()))
                    predicted = self.aiModel.predict(test_x)
                    
                    print('['+self.areaName[counter]+']', predicted[0],end=' ')
                    self.aiResultDict[self.areaName[counter]]=predicted[0]
                    if self.aiResult[counter]:
                        self.canvas.delete(self.aiResult[counter])
                    self.aiResult[counter]=self.canvas.create_text(startCol+3, endRow+12, text=predicted[0], fill="RED", font=('Helvetica 15'), anchor = 'w')
                    
                    counter = counter + 1
                print("",end="\r")
                if self.logInterval != 0:
                    next_datetime = self.lastLogTime + timedelta(seconds=self.logInterval)
                    if now >= next_datetime:
                        self.saveLogFile()
                        self.lastLogTime=now

            self.window.after(500, self.runAI)


def checkImage(image):
    # display image under test
    # return : none
    cv2.imshow('Check Image', image)
    cv2.moveWindow('Check Image', 40+640+20,40)
    cv2.waitKey(displayTime)

 # Change image color to balck and white
def prepareImageM(image,nsize,display=False,color='COLOR'):
    # select color of image in memory
    # return: none
    if color=='GRAY' or color=='BW':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    image = cv2.resize(image, (nsize,nsize))

    if color=='BW':
        (thresh, image) = cv2.threshold(image, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    if display:
        checkImage(image)
    tv = np.array(image.flat)
    return tv

def prepareImageF(file, nsize,display=False,color='COLOR'):
    # select color of image in file
    # return: none
    if color=='COLOR':
        image = cv2.imread(file, cv2.IMREAD_COLOR)
    if color=='GRAY' or color=='BW':
        image = cv2.imread(file, cv2.IMREAD_GRAYSCALE)

    image = cv2.resize(image, (nsize, nsize))   
    if color=='BW':
        (thresh, image) = cv2.threshold(image, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    if display:
        checkImage(image)
    tv = np.array(image.flat)
    return tv

### MAIN ########################################
# Start Main GUI 
#Start Login Page
AppLogin(Tk())
#Start main GUI
App(Tk(), "Automatic data logger"+" "+version,0)
